import openpyxl
from pathlib import Path
from datasource import Repository


def import_responses(repository: Repository):
    excel = Path('ai_responses.xlsx')
    work_book = openpyxl.load_workbook(excel)

    sheet = work_book.active

    id_dictionary = {}

    for i in range(2, sheet.max_row + 1):
        resp = sheet[f'F{i}'].value
        if resp is None:
            break
        resp_id = repository.insert_response(resp)
        id_dictionary[sheet[f'E{i}'].value] = resp_id
    work_book.close()
    import_relations(repository, id_dictionary)


def import_relations(repository: Repository, id_dictionary: dict):
    excel = Path('ai_responses.xlsx')
    work_book = openpyxl.load_workbook(excel)

    sheet = work_book.active

    for i in range(2, sheet.max_row + 1):
        req_id = sheet[f'B{i}'].value
        resp_id = id_dictionary[sheet[f'C{i}'].value]
        repository.insert_request_response(req_id, resp_id)

    work_book.close()


if __name__ == '__main__':
    import_responses(Repository())
